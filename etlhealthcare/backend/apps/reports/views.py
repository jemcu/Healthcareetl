import csv
import io
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiResponse
from drf_spectacular.types import OpenApiTypes
from apps.etl.models import Paciente

FIELDS = [
    "id_paciente", "nombres", "apellidos", "edad", "sexo", "peso", "altura",
    "imc", "imc_clasificacion", "presion_sistolica", "presion_diastolica",
    "frecuencia_cardiaca", "glucosa", "colesterol", "saturacion_oxigeno",
    "temperatura", "fumador", "consumo_alcohol", "antecedentes_familiares",
    "diagnostico_preliminar", "riesgo_enfermedad", "fecha_consulta",
]

@extend_schema(
    tags=["Reports"],
    parameters=[
        OpenApiParameter("tipo", OpenApiTypes.STR, OpenApiParameter.PATH,
                         description="Formato del reporte: csv | xlsx | pdf",
                         enum=["csv", "xlsx", "pdf"]),
    ],
    responses={
        200: OpenApiResponse(description="Archivo descargable en el formato solicitado"),
        400: OpenApiResponse(description="Formato no soportado"),
    },
)
class ReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, tipo):
        qs = Paciente.objects.all()
        if tipo == "csv":
            return self._csv(qs)
        if tipo == "xlsx":
            return self._xlsx(qs)
        if tipo == "pdf":
            return self._pdf(qs)
        return HttpResponse("Formato no soportado", status=400)

    def _csv(self, qs):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="reporte_pacientes.csv"'
        response.write("\ufeff")  # BOM para Excel
        writer = csv.writer(response)
        writer.writerow(FIELDS)
        for p in qs.values_list(*FIELDS):
            writer.writerow(p)
        return response

    def _xlsx(self, qs):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse("openpyxl no instalado", status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pacientes"

        # Cabecera con estilo
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1A3C34")

        for col_num, field in enumerate(FIELDS, 1):
            cell = ws.cell(row=1, column=col_num, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(len(field) + 4, 14)

        # Datos
        for row_num, p in enumerate(qs.values_list(*FIELDS), 2):
            for col_num, value in enumerate(p, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_pacientes.xlsx"'
        return response

    def _pdf(self, qs):
        try:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            return HttpResponse("reportlab no instalado", status=500)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20,
                                topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []

        # Título
        elements.append(Paragraph("Reporte Clínico · HealthAnalytics IPS", styles["Title"]))
        elements.append(Spacer(1, 12))

        # Solo primeros 500 para no saturar el PDF
        pacientes = list(qs.values_list(*FIELDS)[:500])

        # Columnas visibles en PDF (subset para que quepa)
        pdf_fields = ["id_paciente", "nombres", "apellidos", "edad", "sexo",
                      "diagnostico_preliminar", "riesgo_enfermedad", "fecha_consulta"]
        pdf_indices = [FIELDS.index(f) for f in pdf_fields]

        data = [pdf_fields]  # cabecera
        for row in pacientes:
            data.append([str(row[i]) if row[i] is not None else "" for i in pdf_indices])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3C34")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F0E8")]),
            ("GRID",       (0, 0), (-1, -1), 0.3, colors.grey),
            ("ALIGN",      (0, 0), (-1, -1), "LEFT"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="reporte_clinico.pdf"'
        return response
